from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parents[1]
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from nhb_utils import NHB, NHB_TABLES, TABLES, append_manifest, append_registry, ensure_nhb_dirs


ANALYSIS_ID = "nhb_27_method_comparison_figure_source_data"
SCRIPT_NAME = "scripts/nhb_revision/27_compare_methods_and_build_figure_data.py"
DOC_A = Path(r"C:\Users\Gebruiker\Downloads\machine_defined_state_capacity_profiles_native_equations.docx")
DOC_B = Path(r"C:\Users\Gebruiker\Downloads\state_capacity_manuscript_NHB.docx")
OUT_DIR = NHB / "manuscript_comparison"
METHOD_REPORT = OUT_DIR / "method_section_comparison_report.md"
FIGURE_WORKBOOK = OUT_DIR / "state_capacity_manuscript_NHB_proposed_figure_source_data.xlsx"
EXTRACTION_A = OUT_DIR / "machine_defined_state_capacity_profiles_native_equations_methods.txt"
EXTRACTION_B = OUT_DIR / "state_capacity_manuscript_NHB_methods.txt"
FIGURE_CAPTIONS = OUT_DIR / "state_capacity_manuscript_NHB_figure_captions.csv"


@dataclass
class DocSections:
    path: Path
    paragraphs: list[tuple[int, str, str]]
    methods: list[str]
    captions: list[dict[str, str]]


ACTUAL_ANALYSIS = {
    "datasets": "COG-BCI, ds007554, TU Berlin EEG-NIRS, HBN Release 4",
    "events": 289827,
    "architecture_gate": "true vanilla RNN, GRU and LSTM artificial-agent perturbation gate",
    "human_model": "compact GRU profile estimation for human data",
    "state_status": "state is a qualified session/task reliability profile, not a fully validated scalar coordinate",
    "capacity_status": "capacity is the strongest profile, supported by recurrent geometry and TU Berlin load pressure",
    "physiology_status": "EEG/fNIRS/ECG alignment is exploratory and bounded, not direct neural-coordinate evidence",
    "ds007554_status": "ds007554 supervised correctness is reconstructed from push-button timing where needed",
}


def read_doc(path: Path) -> DocSections:
    doc = Document(path)
    paragraphs: list[tuple[int, str, str]] = []
    for i, para in enumerate(doc.paragraphs):
        txt = " ".join(para.text.split())
        if txt:
            style_name = para.style.name if para.style is not None else "NoStyle"
            paragraphs.append((i, style_name, txt))
    methods = extract_methods(paragraphs)
    captions = extract_captions(paragraphs)
    return DocSections(path=path, paragraphs=paragraphs, methods=methods, captions=captions)


def is_heading_style(style: str) -> bool:
    return style.lower().startswith("heading") or style.lower() in {"title", "subtitle"}


def extract_methods(paragraphs: list[tuple[int, str, str]]) -> list[str]:
    start = None
    for idx, (_, style, text) in enumerate(paragraphs):
        normalized = text.strip().lower()
        if normalized in {"methods", "method", "materials and methods", "online methods"}:
            start = idx + 1
            break
    if start is None:
        # Some manuscripts use "Methods" as part of a title. Fall back to first paragraph
        # containing a method heading.
        for idx, (_, style, text) in enumerate(paragraphs):
            if re.match(r"^(methods|method)\b", text.strip(), flags=re.I):
                start = idx + 1
                break
    if start is None:
        return []
    stop_terms = {
        "results",
        "discussion",
        "references",
        "acknowledgements",
        "author contributions",
        "data availability",
        "code availability",
        "extended data",
        "figure legends",
        "figures",
    }
    out: list[str] = []
    for _, style, text in paragraphs[start:]:
        norm = text.strip().lower()
        if norm in stop_terms:
            break
        if is_heading_style(style) and norm in stop_terms:
            break
        if re.match(r"^fig(?:ure)?\.?\s*\d+", text, flags=re.I):
            break
        out.append(text)
    return out


def extract_captions(paragraphs: list[tuple[int, str, str]]) -> list[dict[str, str]]:
    captions: list[dict[str, str]] = []
    for _, style, text in paragraphs:
        if re.match(r"^(figure|fig\.?)\s*\d+", text.strip(), flags=re.I):
            m = re.match(r"^(figure|fig\.?)\s*(\d+)[\.:|\s-]*(.*)$", text.strip(), flags=re.I)
            if m:
                fig_no = m.group(2)
                caption = m.group(3).strip() or text.strip()
            else:
                fig_no = str(len(captions) + 1)
                caption = text.strip()
            captions.append({"figure": f"Figure {fig_no}", "caption": caption, "style": style})
    return captions


def term_present(text: str, patterns: Iterable[str]) -> bool:
    lower = text.lower()
    return any(p.lower() in lower for p in patterns)


def score_methods(methods: list[str]) -> tuple[int, list[str], list[str]]:
    text = "\n".join(methods)
    positives: list[str] = []
    cautions: list[str] = []
    score = 0
    checks = [
        ("Mentions open/public multi-dataset analysis", ["cog-bci", "ds007554", "tu berlin", "hbn"], 2),
        ("Mentions artificial perturbation validation before human claims", ["artificial", "perturbation", "gate"], 2),
        ("Mentions true architecture variants or RNN/GRU/LSTM robustness", ["vanilla rnn", "lstm", "gru"], 2),
        ("Treats state/capacity as profiles or bounded coordinates", ["profile", "profiles"], 2),
        ("Mentions state as reliability/operating profile", ["reliability", "operating", "state"], 1),
        ("Mentions capacity pressure/load validation", ["capacity pressure", "load-by-capacity", "load pressure"], 2),
        ("Mentions physiology as EEG/fNIRS/ECG", ["eeg", "fnirs", "ecg"], 1),
        ("Mentions claim grading/audit/falsification", ["claim", "audit", "falsification"], 2),
    ]
    for label, pats, points in checks:
        if term_present(text, pats):
            positives.append(label)
            score += points
    caution_checks = [
        ("Overstates scalar coordinates if not qualified", ["validated scalar coordinate", "clean scalar coordinate", "coordinate system for human cognition"]),
        ("Overstates direct neural evidence", ["direct neural", "neural coordinate", "neural basis"]),
        ("Describes methods not implemented in current pipeline", ["hierarchical bayesian", "variational inference", "kalman", "particle filter", "source-localized"]),
        ("Uses theoretical equations without matching implemented scripts", ["stochastic differential", "free energy", "active inference"]),
        ("Implies all gates passed cleanly", ["all validation gates passed", "fully recovered state"]),
    ]
    for label, pats in caution_checks:
        if term_present(text, pats):
            cautions.append(label)
            score -= 2
    return score, positives, cautions


def compare_methods(doc_a: DocSections, doc_b: DocSections) -> str:
    score_a, pos_a, caut_a = score_methods(doc_a.methods)
    score_b, pos_b, caut_b = score_methods(doc_b.methods)
    winner = doc_a.path.name if score_a > score_b else doc_b.path.name if score_b > score_a else "tie"
    truth = []
    truth.append("# Method Section Comparison")
    truth.append(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
    truth.append("")
    truth.append("## Actual Implemented Analysis Anchors")
    for key, value in ACTUAL_ANALYSIS.items():
        truth.append(f"- **{key}**: {value}")
    truth.append("")
    truth.append("## Verdict")
    if winner == doc_b.path.name:
        truth.append(
            f"**The second file, `{doc_b.path.name}`, is closer to the actual implemented analysis**, provided its state/capacity language remains bounded as profiles rather than fully validated scalar coordinates."
        )
    elif winner == doc_a.path.name:
        truth.append(
            f"**The first file, `{doc_a.path.name}`, is closer by keyword scoring**, but review the cautions below carefully before using it as the final Methods section."
        )
    else:
        truth.append("The two Methods sections are close by automated scoring; use the cautions below to decide paragraph by paragraph.")
    truth.append("")
    truth.append("## File-by-File Assessment")
    for label, doc, score, positives, cautions in [
        ("File 1", doc_a, score_a, pos_a, caut_a),
        ("File 2", doc_b, score_b, pos_b, caut_b),
    ]:
        truth.append(f"### {label}: `{doc.path.name}`")
        truth.append(f"- Methods paragraphs extracted: {len(doc.methods)}")
        truth.append(f"- Alignment score: {score}")
        truth.append("- Matches actual analysis:")
        truth.extend([f"  - {p}" for p in positives] or ["  - No major matches detected."])
        truth.append("- Cautions / mismatches:")
        truth.extend([f"  - {c}" for c in cautions] or ["  - No major automated cautions detected."])
        truth.append("")
    truth.append("## Practical Recommendation")
    truth.append(
        "Use the second manuscript's Methods as the main base if it states the current results as machine-defined **profiles**, true artificial-agent architecture robustness, qualified state recovery, capacity pressure, and exploratory physiology. Use equation-heavy material from the first file only as formal background or supplementary mathematical notation unless every equation maps directly to an implemented script/output."
    )
    truth.append("")
    truth.append("## Key Phrases That Must Remain in the Final Methods")
    truth.extend(
        [
            "- True artificial agents were trained/evaluated across vanilla RNN, GRU and LSTM families.",
            "- Human profile estimation is based on the compact GRU pipeline unless explicitly stated otherwise.",
            "- State is a session/task operating-reliability profile; scalar state recovery is qualified.",
            "- Capacity is participant-level in the current implementation and has the strongest validation through geometry and load pressure.",
            "- Physiology uses EEG/fNIRS/ECG feature alignment and blocked permutation controls, but does not prove a neural coordinate.",
            "- ds007554 correctness reconstruction is a bounded-data limitation.",
        ]
    )
    return "\n".join(truth) + "\n"


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    sep = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        return pd.read_csv(path, sep=sep)
    except Exception:
        return pd.DataFrame()


def candidate_sources_for_caption(fig_no: str, caption: str) -> list[Path]:
    cap = caption.lower()
    sources: list[Path] = []
    fancy = NHB / "fancy_figures" / "source_data"
    figure_number_map = {
        "Figure 1": [
            NHB_TABLES / "fig1_concept_pipeline_source_data.csv",
            fancy / "fancy_fig1_graphical_abstract_source_data.csv",
            NHB_TABLES / "architecture_perturbation_gate_results.csv",
            NHB_TABLES / "architecture_hybrid_recovery_results.csv",
        ],
        "Figure 2": [
            NHB_TABLES / "fig2_dataset_pipeline_scale_source_data.csv",
            TABLES / "event_counts_by_dataset.csv",
            TABLES / "event_counts_by_subject_task.csv",
            TABLES / "human_projection_dataset_eligibility.csv",
        ],
        "Figure 3": [
            NHB_TABLES / "fig3_artificial_agent_gates_source_data.csv",
            fancy / "fancy_fig3_architecture_robustness_source_data.csv",
            NHB_TABLES / "architecture_perturbation_gate_results.csv",
            NHB_TABLES / "architecture_hybrid_recovery_results.csv",
            NHB_TABLES / "leave_one_architecture_gate_results.csv",
        ],
        "Figure 4": [
            NHB_TABLES / "fig4_prediction_baselines_source_data.csv",
            NHB_TABLES / "incremental_value_model_comparison.csv",
            NHB_TABLES / "incremental_value_delta_metrics.csv",
            TABLES / "ds007554_discovery_model_comparison.csv",
        ],
        "Figure 5": [
            NHB_TABLES / "fig5_capacity_geometry_pressure_source_data.csv",
            fancy / "fancy_fig4_capacity_pressure_surface_source_data.csv",
            NHB_TABLES / "capacity_pressure_models.csv",
            NHB_TABLES / "capacity_pressure_marginal_effects.csv",
            NHB_TABLES / "capacity_variant_validation_summary.csv",
            TABLES / "recurrent_dynamics_state_capacity_tests.csv",
        ],
        "Figure 6": [
            NHB_TABLES / "fig6_state_reliability_source_data.csv",
            fancy / "fancy_fig5_state_reliability_atlas_source_data.csv",
            NHB_TABLES / "state_split_half_reliability.csv",
            NHB_TABLES / "state_bootstrap_reliability.csv",
            NHB_TABLES / "state_capacity_variance_decomposition.csv",
            NHB_TABLES / "state_early_late_model_comparison.csv",
        ],
        "Figure 7": [
            NHB_TABLES / "fig7_physiology_alignment_source_data.csv",
            fancy / "fancy_fig6_physiology_claim_audit_source_data.csv",
            NHB_TABLES / "physiology_robustness_models.csv",
            NHB_TABLES / "physiology_permutation_controls.csv",
            TABLES / "ds007554_neurophys_models.csv",
            TABLES / "cog_bci_validation_models.csv",
        ],
        "Figure 8": [
            NHB_TABLES / "fig8_claim_audit_falsification_source_data.csv",
            NHB_TABLES / "leave_one_dataset_out_validation.csv",
            NHB_TABLES / "leave_one_task_out_validation.csv",
            NHB_TABLES / "profile_shuffle_controls.csv",
            NHB_TABLES / "capacity_cross_task_consistency_controls.csv",
            NHB_TABLES / "feature_taxonomy_enrichment.csv",
            NHB / "audit" / "nhb_final_claim_audit.tsv",
        ],
    }
    if fig_no in figure_number_map:
        sources.extend(figure_number_map[fig_no])
    else:
        # Semantic fallback for future manuscripts with non-standard numbering.
        pass
    mappings = [
        (["graphical", "pipeline", "claim", "audit", "overview"], [fancy / "fancy_fig1_graphical_abstract_source_data.csv", NHB_TABLES / "fig1_concept_pipeline_source_data.csv"]),
        (["state-capacity", "landscape", "profile space", "map"], [fancy / "fancy_fig2_state_capacity_landscape_source_data.csv"]),
        (["architecture", "perturbation", "artificial", "agent", "gate", "hybrid"], [fancy / "fancy_fig3_architecture_robustness_source_data.csv", NHB_TABLES / "architecture_perturbation_gate_results.csv", NHB_TABLES / "architecture_hybrid_recovery_results.csv", NHB_TABLES / "leave_one_architecture_gate_results.csv"]),
        (["capacity pressure", "load", "n-back", "nback", "tu berlin"], [fancy / "fancy_fig4_capacity_pressure_surface_source_data.csv", NHB_TABLES / "capacity_pressure_models.csv", NHB_TABLES / "capacity_pressure_marginal_effects.csv"]),
        (["state reliability", "split", "variance", "icc", "early"], [fancy / "fancy_fig5_state_reliability_atlas_source_data.csv", NHB_TABLES / "state_capacity_variance_decomposition.csv", NHB_TABLES / "state_bootstrap_reliability.csv"]),
        (["physiology", "eeg", "fnirs", "ecg", "neuro"], [fancy / "fancy_fig6_physiology_claim_audit_source_data.csv", NHB_TABLES / "physiology_robustness_models.csv", NHB_TABLES / "physiology_permutation_controls.csv"]),
        (["prediction", "baseline", "incremental"], [NHB_TABLES / "fig4_prediction_baselines_source_data.csv", NHB_TABLES / "incremental_value_model_comparison.csv"]),
        (["dataset", "scale", "sample", "event"], [NHB_TABLES / "fig2_dataset_pipeline_scale_source_data.csv", TABLES / "event_counts_by_dataset.csv"]),
    ]
    if fig_no not in figure_number_map:
        for keywords, paths in mappings:
            if any(k in cap for k in keywords):
                sources.extend(paths)
    # Deduplicate while preserving order.
    out: list[Path] = []
    seen = set()
    for p in sources:
        if p.exists() and p not in seen:
            out.append(p)
            seen.add(p)
    return out


def safe_sheet_name(name: str, used: set[str]) -> str:
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)[:31] or "sheet"
    base = name
    i = 1
    while name in used:
        suffix = f"_{i}"
        name = f"{base[:31-len(suffix)]}{suffix}"
        i += 1
    used.add(name)
    return name


def add_dataframe_sheet(wb: Workbook, name: str, df: pd.DataFrame, used: set[str], max_rows: int = 100000) -> str:
    ws = wb.create_sheet(safe_sheet_name(name, used))
    if df.empty:
        ws.append(["note"])
        ws.append(["No source rows available."])
        return ws.title
    out = df.head(max_rows).copy()
    ws.append(list(out.columns))
    for row in out.itertuples(index=False, name=None):
        ws.append([None if pd.isna(v) else v for v in row])
    format_sheet(ws)
    if len(df) > max_rows:
        ws.cell(row=ws.max_row + 2, column=1, value=f"NOTE: source truncated from {len(df)} to {max_rows} rows for Excel size.")
    return ws.title


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="F4F6F9")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="14213D")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx in range(1, min(ws.max_column, 30) + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter][: min(ws.max_row, 200)]:
            if cell.value is not None:
                max_len = min(max(max_len, len(str(cell.value)) + 2), 55)
        ws.column_dimensions[letter].width = max_len
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 1000)):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def build_figure_workbook(doc_b: DocSections) -> tuple[Path, pd.DataFrame]:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    index_rows: list[dict[str, object]] = []
    source_map_rows: list[dict[str, object]] = []
    captions = doc_b.captions
    if not captions:
        # If no explicit captions are found, use current manuscript-facing figure plan.
        captions = [
            {"figure": "Figure 1", "caption": "Study logic and claim governance."},
            {"figure": "Figure 2", "caption": "Human state-capacity landscape."},
            {"figure": "Figure 3", "caption": "Architecture robustness."},
            {"figure": "Figure 4", "caption": "Capacity pressure under N-back load."},
            {"figure": "Figure 5", "caption": "State reliability atlas."},
            {"figure": "Figure 6", "caption": "Physiology and bounded claims."},
        ]
    for cap in captions:
        fig = cap["figure"]
        caption = cap["caption"]
        sources = candidate_sources_for_caption(fig, caption)
        if not sources:
            index_rows.append({"figure": fig, "caption": caption, "sheet": "", "source_files": "", "status": "no_matching_source_found"})
            continue
        merged_parts = []
        for src in sources:
            df = read_csv(src)
            if df.empty:
                continue
            df = df.copy()
            df.insert(0, "source_file", src.name)
            df.insert(1, "source_path", str(src.relative_to(ROOT)))
            merged_parts.append(df)
            source_map_rows.append({"figure": fig, "caption": caption, "source_file": src.name, "source_path": str(src.relative_to(ROOT)), "rows": len(df), "columns": len(df.columns)})
        merged = pd.concat(merged_parts, ignore_index=True, sort=False) if merged_parts else pd.DataFrame()
        sheet = add_dataframe_sheet(wb, f"{fig.replace(' ', '')}_data", merged, used)
        index_rows.append({"figure": fig, "caption": caption, "sheet": sheet, "source_files": "; ".join(p.name for p in sources), "status": "ok" if not merged.empty else "empty_sources"})
    # Add index and source map first by creating then moving to front.
    index_df = pd.DataFrame(index_rows)
    map_df = pd.DataFrame(source_map_rows)
    idx_sheet = add_dataframe_sheet(wb, "INDEX", index_df, used)
    map_sheet = add_dataframe_sheet(wb, "SOURCE_MAP", map_df, used)
    wb._sheets = [wb[idx_sheet], wb[map_sheet]] + [s for s in wb.worksheets if s.title not in {idx_sheet, map_sheet}]
    FIGURE_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(FIGURE_WORKBOOK)
    return FIGURE_WORKBOOK, index_df


def main() -> None:
    ensure_nhb_dirs()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    started = datetime.now(timezone.utc).isoformat()
    doc_a = read_doc(DOC_A)
    doc_b = read_doc(DOC_B)
    EXTRACTION_A.write_text("\n\n".join(doc_a.methods), encoding="utf-8")
    EXTRACTION_B.write_text("\n\n".join(doc_b.methods), encoding="utf-8")
    captions_df = pd.DataFrame(doc_b.captions)
    captions_df.to_csv(FIGURE_CAPTIONS, index=False)
    report = compare_methods(doc_a, doc_b)
    METHOD_REPORT.write_text(report, encoding="utf-8")
    workbook, index_df = build_figure_workbook(doc_b)
    outputs = [METHOD_REPORT, FIGURE_WORKBOOK, EXTRACTION_A, EXTRACTION_B, FIGURE_CAPTIONS]
    append_manifest(ANALYSIS_ID, SCRIPT_NAME, outputs)
    append_registry(ANALYSIS_ID, SCRIPT_NAME, started, outputs, notes=f"Compared Methods sections and generated figure source-data workbook with {len(index_df)} figure entries.")
    print(f"Wrote {METHOD_REPORT}")
    print(f"Wrote {FIGURE_WORKBOOK}")


if __name__ == "__main__":
    main()
